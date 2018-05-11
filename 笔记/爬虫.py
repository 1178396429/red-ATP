不要用正则表达式来解析 HTML
KfrEW7jTMOec

打开网页  webbrowser

import webbrowser
for i in range(10):
    webbrowser.open('http://inventwithpython.com/')





requests--------------------------------------------------------------

检查错误

res = requests.get('http://inventwithpython.com/page_that_does_not_exist')
res.raise_for_status()



原始响应内容
Response.iter_content是可迭代对象

>>> r = requests.get('https://github.com/timeline.json', stream=True)
>>> r.raw
<requests.packages.urllib3.response.HTTPResponse object at 0x101194810>
>>> r.raw.read(10)
'\x1f\x8b\x08\x00\x00\x00\x00\x00\x00\x03'

with open(filename, 'wb') as fd:
    for chunk in r.iter_content(chunk_size):
        fd.write(chunk)






selenium--------------------------------------------------------------------------


from selenium import webdriver
browser = webdriver.Firefox()
browser.get('https://www.baidu.com/')

点击事件
browser.get('https://www.hao123.com/')
link = browser.find_element_by_link_text('爱 奇 艺')
link.click()

browser.back()       点击“返回”按钮。 
browser.forward()    点击“前进”按钮。
browser.refresh()    点击“刷新”按钮。
browser.quit()       点击“关闭窗口”按钮








Excel---------------------------------------------------------------
wb.get_sheet_names(列表) 获得表名
sheet = wb.get_sheet_by_name('Sheet3')  获得表3的名字
sheet.title  表格的title

import openpyxl,os
os.chdir("Z:\\e\\")
wb = openpyxl.load_workbook('51.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
sheet['A1'].value

 sheet.cell(row=1, column=2).value
 'Apples'
 for i in range(1, 8, 2):
 	print(i, sheet.cell(row=i, column=2).value)



创建保存
from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws['A1'] = 42
ws.append([1, 2, 3])
import datetime
ws['A2'] = datetime.datetime.now()
wb.save("sample.xlsx")

循环行和列
import openpyxl,os
os.chdir(".")
wb = openpyxl.load_workbook('51.xlsx')
sheet = wb.get_sheet_by_name('s')
sheet['A1'] = '123456'
for i in range(1, 8):
 	print(i, sheet.cell(row=i, column=2).value)


打印A1到C3
 import openpyxl,os
 os.chdir(".")
 wb = openpyxl.load_workbook('51.xlsx')
 sheet = wb.get_sheet_by_name('Sheet1')
 for rowOfCellObjects in sheet['A1':'C3']:
     for cellObj in rowOfCellObjects:
         print(cellObj.coordinate, cellObj.value)
 print('--- END OF ROW ---')

打印表中数据
 for row in wb['sheet'].rows:
         print( *[ cell.value for cell in row ])

 for row in wb['sheet'].values:
         print(*row)


openpyxl

Workbook是对工作簿的抽象，Worksheet是对表格的抽象，Cell是对单元格的抽象


Workbook属性

	wb.active      ：获取当前活跃的Worksheet
	wb.worksheets  ：以列表的形式返回所有的Worksheet(表格)
	wb.read_only   ：判断是否以read_only模式打开Excel文档
	wb.encoding    ：获取文档的字符集编码
	wb.properties  ：获取文档的元数据，如标题，创建者，创建日期等
	wb.sheetnames  ：获取工作簿中的表（列表）

	 
	方法
    
	wb = openpyxl.load_workbook('51.xlsx')
	ws = wb['sheet']
	print(ws['A1'].value)
	get_sheet_names 废弃  ---->换成  wb.sheetnames
	get_sheet_by_name    ---->       wb['sheet']

	get_active_sheet    ：获取活跃的表格(新版建议通过active属性获取)
	remove_sheet        ：删除一个表格
	create_sheet        ：创建一个空的表格
	copy_worksheet      ：在Workbook内拷贝表格
	


Worksheet对象 属性
	title              ：表格的标题                                                                                                                                                            
	dimensions         ：表格的大小，这里的大小是指含有数据的表格的大小，即 左上角的坐标:右下角的坐标
	max_row            ：表格的最大行                                                                                                                                                
	min_row            ：表格的最小行                                                                                                                                                
	max_column         ：表格的最大列                                                                                                                                    
	min_column         ：表格的最小列                                                                                                                                    
	rows               ：按行获取单元格(Cell对象) - 生成器                                                                                                
	columns            ：按列获取单元格(Cell对象) - 生成器                                                                                    
	freeze_panes       ：冻结窗格                                                                                                                                    
	values             ：按行获取表格的内容(数据)  - 生成器                                                                                            
	
	方法
	iter_rows：按行获取所有单元格，内置属性有(min_row,max_row,min_col,max_col)
	iter_columns：按列获取所有的单元格
	append：在表格末尾添加数据
	merged_cells：合并多个单元格
	unmerged_cells：移除合并的单元格


cell

	row：单元格所在的行
	column：单元格坐在的列
	value：单元格的值
	coordinate：单元格的坐标	
 print(wb['sheet'].cell(row=1,column=2).coordinate)   B1
 print(wb['sheet'].cell(row=1,column=2).value)        ABC 
 print(wb['sheet'].cell(row=1,column=2).row)          B
 print(wb['sheet'].cell(row=1,column=2).column)       1
 